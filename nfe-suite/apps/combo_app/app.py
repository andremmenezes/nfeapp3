# nfe-suite/apps/combo_app/app.py
# XML: gera o MESMO Excel do xml_app (colunas e formatação).
# PDF: tenta extractor externo; se não houver, usa PDFMiner (texto) e, por fim, PyPDF.
#      Varredura robusta por chaves 44 dígitos. Informa se o PDF parece ser imagem.

import io
import os
import sys
import re
import zipfile
import tempfile
from pathlib import Path
from datetime import datetime
from io import BytesIO
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# =========================
# Config da página
# =========================
st.set_page_config(page_title="NFe Suite", layout="wide")
st.title("NFe Suite – Upload e Processamento")

# =========================
# Tentativa de usar extractor externo (opcional)
# =========================
EXTERNAL_EXTRACTOR = None
try:
    HERE = Path(__file__).resolve()
    PDF_APP_DIR = (HERE.parent.parent / "pdf_app").resolve()  # .../apps/pdf_app
    if PDF_APP_DIR.exists():
        if str(PDF_APP_DIR) not in sys.path:
            sys.path.insert(0, str(PDF_APP_DIR))
        from extractor import extrair_chaves_de_pdf as EXTERNAL_EXTRACTOR  # type: ignore
except Exception:
    EXTERNAL_EXTRACTOR = None

# =========================
# Helpers gerais de arquivos
# =========================
def save_uploaded_files(files, dest_dir: Path) -> list[Path]:
    dest_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    for f in files:
        p = dest_dir / f.name
        with open(p, "wb") as out:
            out.write(f.getbuffer())
        saved.append(p)
    return saved

def extract_zip_to(zip_bytes: bytes, dest_dir: Path) -> list[Path]:
    dest_dir.mkdir(parents=True, exist_ok=True)
    saved = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for member in zf.infolist():
            if member.is_dir():
                continue
            target = dest_dir / Path(member.filename).name
            with zf.open(member, "r") as src, open(target, "wb") as dst:
                dst.write(src.read())
            saved.append(target)
    return saved

# =========================
# Helpers de Excel (XML/PDF)
# =========================
def df_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in sheets.items():
            safe = (name or "Planilha")[:31]
            (df if not df.empty else pd.DataFrame()).to_excel(writer, index=False, sheet_name=safe)
    buf.seek(0)
    return buf.getvalue()

def excel_filename(prefix: str = "resultado") -> str:
    return f"{prefix}-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"

# =========================
# === Funções do xml_app (mantém MESMO Excel) ===
# =========================
def get_text_or_zero(elem):
    return elem.text if elem is not None and elem.text else "0"

def extract_info_from_xml(file, result_list):
    tree = ET.parse(file)
    root = tree.getroot()
    ns = {"ns": "http://www.portalfiscal.inf.br/nfe"}

    nfe = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:nNF", ns)
    serie = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:serie", ns)
    nat_operacao = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:natOp", ns)

    data_saida_entrada = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:dhSaiEnt", ns)
    if data_saida_entrada is not None and data_saida_entrada.text:
        data_saida_entrada = f"{data_saida_entrada.text[8:10]}/{data_saida_entrada.text[5:7]}/{data_saida_entrada.text[0:4]}"
    else:
        data_saida_entrada = ""

    data_emissao = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:dhEmi", ns)
    if data_emissao is not None and data_emissao.text:
        data_emissao = f"{data_emissao.text[8:10]}/{data_emissao.text[5:7]}/{data_emissao.text[0:4]}"
    else:
        data_emissao = ""

    valor_frete = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vFrete", ns)
    valor_desc = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vDesc", ns)
    valor_outro = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vOutro", ns)
    valor_tot_nota = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vNF", ns)
    valor_tot_prod = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vProd", ns)

    chave = root.find("./ns:protNFe/ns:infProt/ns:chNFe", ns)
    cnpj_emitente = root.find("./ns:NFe/ns:infNFe/ns:emit/ns:CNPJ", ns)
    nome_emitente = root.find("./ns:NFe/ns:infNFe/ns:emit/ns:xNome", ns)

    item_num = 1
    for item in root.findall("./ns:NFe/ns:infNFe/ns:det", ns):
        cod = item.find(".ns:prod/ns:cProd", ns)
        descricao = item.find(".ns:prod/ns:xProd", ns)
        unidade = item.find(".ns:prod/ns:uCom", ns)
        quantidade = item.find(".ns:prod/ns:qCom", ns)
        valor_unit = item.find(".ns:prod/ns:vUnCom", ns)
        desconto_item = item.find(".ns:prod/ns:vDesc", ns)
        desconto_item = desconto_item.text if desconto_item is not None and desconto_item.text else ""
        valor_total_item = item.find(".ns:prod/ns:vProd", ns)
        icms_percent = item.find(".ns:imposto/ns:ICMS/ns:ICMS00/ns:pICMS", ns)
        icms_valor = item.find(".ns:imposto/ns:ICMS/ns:ICMS00/ns:vICMS", ns)
        ipi_percent = item.find(".ns:imposto/ns:IPI/ns:IPITrib/ns:pIPI", ns)
        ipi_valor = item.find(".ns:imposto/ns:IPI/ns:IPITrib/ns:vIPI", ns)

        linha = [
            nfe.text, serie.text, nat_operacao.text, data_emissao, data_saida_entrada,
            valor_frete.text, chave.text, cnpj_emitente.text, nome_emitente.text,
            valor_tot_nota.text, valor_tot_prod.text, valor_desc.text, valor_outro.text,
            item_num, cod.text, descricao.text, unidade.text, quantidade.text,
            valor_unit.text, desconto_item, valor_total_item.text,
            get_text_or_zero(icms_percent), get_text_or_zero(icms_valor),
            get_text_or_zero(ipi_percent), get_text_or_zero(ipi_valor),
        ]
        result_list.append([str(i) for i in linha])
        item_num += 1

def format_excel(df: pd.DataFrame) -> BytesIO:
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    formatos = {
        "Valor do Frete": "R$ #,##0.00",
        "Valor Total da Nota": "R$ #,##0.00",
        "Valor Total dos Produtos": "R$ #,##0.00",
        "Descontos Aplicados": "R$ #,##0.00",
        "Outras Despesas Acessórias": "R$ #,##0.00",
        "Valor Unitário": "R$ #,##0.00",
        "Desconto": "R$ #,##0.00",
        "Valor Total do Item": "R$ #,##0.00",
        "ICMS (valor)": "R$ #,##0.00",
        "IPI (valor)": "R$ #,##0.00",
        "Data de emissão": "DD/MM/YYYY",
        "Data de Saída/Entrada": "DD/MM/YYYY",
        "ICMS(%)": "0.00%",
        "IPI(%)": "0.00%",
        "Quantidade": "0.00",
        "NFe": "@", "Série": "@", "Chave": "@", "CNPJ do Emitente": "@", "Cód. Produto": "@",
    }

    cab = {cell.value.strip(): i + 1 for i, cell in enumerate(ws[1]) if cell.value}
    for nome, fmt in formatos.items():
        if nome in cab:
            col = cab[nome]
            for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                cell = row[0]
                try:
                    val = str(cell.value or "").replace("R$", "").replace("%", "").replace(",", ".").strip()
                    if val == "":
                        continue
                    if fmt == "0":
                        cell.value = int(float(val))
                    elif fmt.endswith("%"):
                        cell.value = float(val) / 100
                    elif fmt == "@":
                        cell.value = str(cell.value).strip()
                    else:
                        cell.value = float(val)
                    cell.number_format = fmt
                    if nome == "CNPJ do Emitente":
                        cnpj = "".join(filter(str.isdigit, str(cell.value)))
                        if len(cnpj) == 14:
                            cell.value = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                except:
                    pass

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# =========================
# PDF: extração de texto com PDFMiner e PyPDF + busca robusta
# =========================
# PDFMiner (puro Python, melhor para texto)
try:
    from pdfminer.high_level import extract_text as pdfminer_extract_text
    _PDFMINER_OK = True
except Exception:
    _PDFMINER_OK = False
    pdfminer_extract_text = None  # type: ignore

# PyPDF (puro Python, alternativa)
try:
    from pypdf import PdfReader
    _PYPDF_OK = True
except Exception:
    _PYPDF_OK = False
    PdfReader = None  # type: ignore

# regex que aceita separadores entre dígitos
_CHUNK_RE = re.compile(r"(\d[\d\.\-\s]{40,80}\d)")
_ONLY_DIGITS = re.compile(r"\D")

def _normalize_digits(s: str) -> str:
    return _ONLY_DIGITS.sub("", s or "")

def _slide_44(line_digits: str) -> list[str]:
    """varre a sequência de dígitos e extrai todos os blocos de 44 dígitos (sobrepostos)"""
    out = []
    n = len(line_digits)
    for i in range(0, max(0, n - 43)):
        blk = line_digits[i:i+44]
        if len(blk) == 44 and blk.isdigit():
            out.append(blk)
    return out

def extract_keys_with_pdfminer(path: str) -> tuple[list[str], str]:
    if not _PDFMINER_OK or pdfminer_extract_text is None:
        return [], ""
    try:
        text = pdfminer_extract_text(path) or ""
    except Exception:
        return [], ""
    keys: list[str] = []
    # por linha melhora a precisão
    for raw_line in text.splitlines():
        # 1) pega blocos com separadores
        for m in _CHUNK_RE.finditer(raw_line):
            digits = _normalize_digits(m.group(1))
            if len(digits) == 44:
                keys.append(digits)
            elif len(digits) > 44:
                keys.extend(_slide_44(digits))
        # 2) varredura direta na linha (se tiver só dígitos)
        digits_line = _normalize_digits(raw_line)
        if len(digits_line) >= 44:
            keys.extend(_slide_44(digits_line))
    return keys, text

def extract_keys_with_pypdf(path: str) -> tuple[list[str], str]:
    if not _PYPDF_OK or PdfReader is None:
        return [], ""
    try:
        reader = PdfReader(path)
    except Exception:
        return [], ""
    keys: list[str] = []
    txt_all = []
    for pg in reader.pages:
        t = pg.extract_text() or ""
        txt_all.append(t)
        for m in _CHUNK_RE.finditer(t):
            digits = _normalize_digits(m.group(1))
            if len(digits) == 44:
                keys.append(digits)
            elif len(digits) > 44:
                keys.extend(_slide_44(digits))
        digits_line = _normalize_digits(t)
        if len(digits_line) >= 44:
            keys.extend(_slide_44(digits_line))
    return keys, "\n".join(txt_all)

def processar_pdfs(paths: list[Path]):
    """
    Ordem:
      1) extractor externo (se existir)
      2) PDFMiner (texto)
      3) PyPDF (texto)
    Retorna df_resumo (por arquivo) e df_chaves (1 por chave)
    """
    linhas = []
    resumo = []

    for p in sorted(paths, key=lambda x: x.name.lower()):
        try:
            chaves: list[str] = []
            txt_used = ""

            if EXTERNAL_EXTRACTOR is not None:
                # seu extractor retorna (chaves, outras)
                try:
                    ext_keys, ext_others = EXTERNAL_EXTRACTOR(str(p))
                    chaves.extend(ext_keys or [])
                    txt_used = "external"
                except Exception:
                    pass

            if not chaves:
                # tenta PDFMiner
                keys_miner, text_miner = extract_keys_with_pdfminer(str(p))
                chaves.extend(keys_miner or [])
                if keys_miner:
                    txt_used = "pdfminer"
                elif text_miner:
                    txt_used = "pdfminer_no_keys"

            if not chaves:
                # tenta PyPDF
                keys_pypdf, text_pypdf = extract_keys_with_pypdf(str(p))
                chaves.extend(keys_pypdf or [])
                if keys_pypdf:
                    txt_used = "pypdf"
                elif text_pypdf:
                    txt_used = "pypdf_no_keys"

            # dedup preservando ordem
            seen = set()
            chaves = [c for c in chaves if c.isdigit() and (c not in seen and not seen.add(c))]

            for c in chaves:
                linhas.append({"arquivo": p.name, "chave_44": c})

            if chaves:
                resumo.append({
                    "arquivo": p.name,
                    "qtd_chaves_44": len(chaves),
                    "chaves_44": ", ".join(chaves),
                    "metodo": txt_used
                })
            else:
                # sem texto => provavelmente PDF escaneado (imagem)
                resumo.append({
                    "arquivo": p.name,
                    "qtd_chaves_44": 0,
                    "chaves_44": "",
                    "metodo": "sem_texto? (provável PDF escaneado; precisa OCR/leitor de códigos)"
                })

        except Exception as e:
            resumo.append({
                "arquivo": p.name,
                "qtd_chaves_44": 0,
                "chaves_44": "",
                "metodo": f"ERRO: {e}"
            })

    df_chaves = pd.DataFrame(linhas).drop_duplicates().reset_index(drop=True)
    df_resumo = pd.DataFrame(resumo).sort_values("arquivo").reset_index(drop=True)
    return df_resumo, df_chaves

# =========================
# Pasta temporária por sessão
# =========================
session_tmp = Path(tempfile.gettempdir()) / f"nfe_suite_{os.getpid()}"
session_tmp.mkdir(parents=True, exist_ok=True)

# =========================
# Abas
# =========================
tab_xml, tab_pdf, tab_zip = st.tabs(["XML (múltiplos)", "PDF (múltiplos)", "ZIP/Lote"])

# --------- Aba: XML ----------
with tab_xml:
    st.subheader("Enviar XMLs")
    xml_files = st.file_uploader(
        "Selecione um ou mais arquivos .xml",
        type=["xml"],
        accept_multiple_files=True,
        help="Arraste e solte vários arquivos."
    )

    if st.button("Processar XMLs", disabled=not xml_files):
        dest = session_tmp / "xml_uploads"
        paths = save_uploaded_files(xml_files, dest)
        st.success(f"{len(paths)} XML(s) recebidos.")

        colunas = [
            "NFe", "Série", "Natureza da Operação", "Data de emissão", "Data de Saída/Entrada",
            "Valor do Frete", "Chave", "CNPJ do Emitente", "Nome do Emitente",
            "Valor Total da Nota", "Valor Total dos Produtos", "Descontos Aplicados",
            "Outras Despesas Acessórias", "Nº Item na Nota", "Cód. Produto", "Descrição",
            "Unidade de Medida", "Quantidade", "Valor Unitário", "Desconto", "Valor Total do Item",
            "ICMS(%)", "ICMS (valor)", "IPI(%)", "IPI (valor)",
        ]
        notas: list[list[str]] = []
        for p in paths:
            extract_info_from_xml(p, notas)
        df_xml = pd.DataFrame(columns=colunas, data=notas)

        excel_bytes = format_excel(df_xml)
        st.success("✅ XMLs processados!")
        st.download_button(
            label="📥 Baixar Excel (XMLs)",
            data=excel_bytes.getvalue(),
            file_name=f"NotasFiscais-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.dataframe(df_xml.head(50), use_container_width=True)

# --------- Aba: PDF ----------
with tab_pdf:
    st.subheader("Enviar PDFs")
    pdf_files = st.file_uploader(
        "Selecione um ou mais PDFs",
        type=["pdf"],
        accept_multiple_files=True,
        help="O app tenta extractor externo, depois PDFMiner e PyPDF. Se o PDF for escaneado (imagem), será necessário OCR/leitor de código."
    )

    if st.button("Processar PDFs", disabled=not pdf_files):
        dest = session_tmp / "pdf_uploads"
        paths = save_uploaded_files(pdf_files, dest)
        st.success(f"{len(paths)} PDF(s) recebidos.")

        df_resumo, df_chaves = processar_pdfs(paths)

        st.subheader("Resumo por arquivo")
        st.dataframe(df_resumo, use_container_width=True)

        st.subheader("Linhas por chave (deduplicadas)")
        st.dataframe(df_chaves, use_container_width=True)

        xlsx_bytes = df_to_excel_bytes({"Resumo_PDF": df_resumo, "Chaves_PDF": df_chaves})
        st.download_button(
            "📥 Baixar Excel (PDFs)",
            data=xlsx_bytes,
            file_name=excel_filename("pdfs"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.download_button(
            "Baixar chaves (CSV)",
            data=df_chaves.to_csv(index=False).encode("utf-8"),
            file_name="chaves_por_linha.csv",
            mime="text/csv",
            use_container_width=True
        )

# --------- Aba: ZIP/Lote ----------
with tab_zip:
    st.subheader("Enviar ZIP com lote (XMLs e/ou PDFs)")
    zip_file = st.file_uploader("Selecione um .zip", type=["zip"])

    if st.button("Processar ZIP", disabled=not zip_file):
        dest = session_tmp / "zip_extract"
        paths = extract_zip_to(zip_file.getvalue(), dest)
        st.success(f"{len(paths)} arquivo(s) extraído(s).")

        xmls = [p for p in paths if p.suffix.lower() == ".xml"]
        pdfs = [p for p in paths if p.suffix.lower() == ".pdf"]

        sheets: dict[str, pd.DataFrame] = {}

        if xmls:
            colunas = [
                "NFe", "Série", "Natureza da Operação", "Data de emissão", "Data de Saída/Entrada",
                "Valor do Frete", "Chave", "CNPJ do Emitente", "Nome do Emitente",
                "Valor Total da Nota", "Valor Total dos Produtos", "Descontos Aplicados",
                "Outras Despesas Acessórias", "Nº Item na Nota", "Cód. Produto", "Descrição",
                "Unidade de Medida", "Quantidade", "Valor Unitário", "Desconto", "Valor Total do Item",
                "ICMS(%)", "ICMS (valor)", "IPI(%)", "IPI (valor)",
            ]
            notas_zip: list[list[str]] = []
            for p in xmls:
                extract_info_from_xml(p, notas_zip)
            df_xml_zip = pd.DataFrame(columns=colunas, data=notas_zip)
            sheets["XMLs"] = df_xml_zip
            st.write("Prévia XMLs do ZIP")
            st.dataframe(df_xml_zip.head(50), use_container_width=True)

        if pdfs:
            df_resumo_zip, df_chaves_zip = processar_pdfs(pdfs)
            sheets["Resumo_PDF"] = df_resumo_zip
            sheets["Chaves_PDF"] = df_chaves_zip
            st.write("Prévia PDFs do ZIP")
            st.dataframe(df_resumo_zip.head(50), use_container_width=True)

        if sheets:
            xlsx_bytes = df_to_excel_bytes(sheets)
            st.download_button(
                label="📥 Baixar Excel consolidado (ZIP/Lote)",
                data=xlsx_bytes,
                file_name=excel_filename("lote"),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        else:
            st.info("Nenhum XML/PDF válido encontrado no ZIP.")

