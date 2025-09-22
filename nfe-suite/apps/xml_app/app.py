import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from io import BytesIO
from openpyxl import load_workbook

# Função auxiliar
def get_text_or_zero(elem):
    return elem.text if elem is not None and elem.text else "0"

# Função para extrair dados de um único XML
def extract_info_from_xml(file, result_list):
    tree = ET.parse(file)
    root = tree.getroot()
    ns = {"ns": "http://www.portalfiscal.inf.br/nfe"}

    nfe = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:nNF", ns)
    serie = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:serie", ns)
    nat_operacao = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:natOp", ns)

    data_saida_entrada = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:dhSaiEnt", ns)
    if data_saida_entrada is not None and data_saida_entrada.text:
        data_saida_entrada = f"{data_saida_entrada.text[8:10]}/{data_saida_entrada.text[5:7]}/{data_saida_entrada.text[0:4]}"
    else:
        data_saida_entrada = ""

    data_emissao = root.find("./ns:NFe/ns:infNFe/ns:ide/ns:dhEmi", ns)
    if data_emissao is not None and data_emissao.text:
        data_emissao = f"{data_emissao.text[8:10]}/{data_emissao.text[5:7]}/{data_emissao.text[0:4]}"
    else:
        data_emissao = ""

    valor_frete = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vFrete", ns)
    valor_desc = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vDesc", ns)
    valor_outro = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vOutro", ns)
    valor_tot_nota = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vNF", ns)
    valor_tot_prod = root.find("./ns:NFe/ns:infNFe/ns:total/ns:ICMSTot/ns:vProd", ns)

    chave = root.find("./ns:protNFe/ns:infProt/ns:chNFe", ns)
    cnpj_emitente = root.find("./ns:NFe/ns:infNFe/ns:emit/ns:CNPJ", ns)
    nome_emitente = root.find("./ns:NFe/ns:infNFe/ns:emit/ns:xNome", ns)

    item_num = 1
    for item in root.findall("./ns:NFe/ns:infNFe/ns:det", ns):
        cod = item.find(".ns:prod/ns:cProd", ns)
        descricao = item.find(".ns:prod/ns:xProd", ns)
        unidade = item.find(".ns:prod/ns:uCom", ns)
        quantidade = item.find(".ns:prod/ns:qCom", ns)
        valor_unit = item.find(".ns:prod/ns:vUnCom", ns)
        desconto_item = item.find(".ns:prod/ns:vDesc", ns)
        desconto_item = desconto_item.text if desconto_item is not None and desconto_item.text else ""
        valor_total_item = item.find(".ns:prod/ns:vProd", ns)
        icms_percent = item.find(".ns:imposto/ns:ICMS/ns:ICMS00/ns:pICMS", ns)
        icms_valor = item.find(".ns:imposto/ns:ICMS/ns:ICMS00/ns:vICMS", ns)
        ipi_percent = item.find(".ns:imposto/ns:IPI/ns:IPITrib/ns:pIPI", ns)
        ipi_valor = item.find(".ns:imposto/ns:IPI/ns:IPITrib/ns:vIPI", ns)

        linha = [
            nfe.text, serie.text, nat_operacao.text, data_emissao, data_saida_entrada,
            valor_frete.text, chave.text, cnpj_emitente.text, nome_emitente.text,
            valor_tot_nota.text, valor_tot_prod.text, valor_desc.text, valor_outro.text,
            item_num, cod.text, descricao.text, unidade.text, quantidade.text,
            valor_unit.text, desconto_item, valor_total_item.text,
            get_text_or_zero(icms_percent), get_text_or_zero(icms_valor),
            get_text_or_zero(ipi_percent), get_text_or_zero(ipi_valor),
        ]

        result_list.append([str(i) for i in linha])
        item_num += 1

# Função para formatar o Excel
def format_excel(df):
    output = BytesIO()
    df.to_excel(output, index=False, engine="openpyxl")
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    formatos_colunas = {
        "Valor do Frete": "R$ #,##0.00",
        "Valor Total da Nota": "R$ #,##0.00",
        "Valor Total dos Produtos": "R$ #,##0.00",
        "Descontos Aplicados": "R$ #,##0.00",
        "Outras Despesas Acessórias": "R$ #,##0.00",
        "Valor Unitário": "R$ #,##0.00",
        "Desconto": "R$ #,##0.00",
        "Valor Total do Item": "R$ #,##0.00",
        "ICMS (valor)": "R$ #,##0.00",
        "IPI (valor)": "R$ #,##0.00",
        "Data de emissão": "DD/MM/YYYY",
        "Data de Saída/Entrada": "DD/MM/YYYY",
        "ICMS(%)": "0.00%",
        "IPI(%)": "0.00%",
        "Quantidade": "0.00",
        "NFe": "@",
        "Série": "@",
        "Chave": "@",
        "CNPJ do Emitente": "@",
        "Cód. Produto": "@",
    }

    cabecalho = {cell.value.strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}

    for nome_coluna, formato in formatos_colunas.items():
        if nome_coluna in cabecalho:
            col_idx = cabecalho[nome_coluna]
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                try:
                    valor = str(cell.value or "").replace("R$", "").replace("%", "").replace(",", ".").strip()
                    if valor == "":
                        continue
                    if formato == "0":
                        cell.value = int(float(valor))
                    elif formato.endswith("%"):
                        cell.value = float(valor) / 100
                    elif formato == "@":
                        cell.value = str(cell.value).strip()
                    else:
                        cell.value = float(valor)
                    cell.number_format = formato
                    if nome_coluna == "CNPJ do Emitente":
                        cnpj = "".join(filter(str.isdigit, str(cell.value)))
                        if len(cnpj) == 14:
                            cell.value = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                except:
                    pass

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# Interface Streamlit
st.set_page_config(page_title="Leitor de NF-e", layout="wide")
st.title("📄 Leitor de XML de NF-e")
st.write("Faça upload de um ou mais arquivos XML para gerar a planilha Excel formatada.")

uploaded_files = st.file_uploader("Selecione os arquivos XML", type=["xml"], accept_multiple_files=True)

if uploaded_files:
    notas = []
    colunas = [
        "NFe", "Série", "Natureza da Operação", "Data de emissão", "Data de Saída/Entrada",
        "Valor do Frete", "Chave", "CNPJ do Emitente", "Nome do Emitente",
        "Valor Total da Nota", "Valor Total dos Produtos", "Descontos Aplicados",
        "Outras Despesas Acessórias", "Nº Item na Nota", "Cód. Produto", "Descrição",
        "Unidade de Medida", "Quantidade", "Valor Unitário", "Desconto", "Valor Total do Item",
        "ICMS(%)", "ICMS (valor)", "IPI(%)", "IPI (valor)",
    ]

    for file in uploaded_files:
        extract_info_from_xml(file, notas)

    df = pd.DataFrame(columns=colunas, data=notas)
    excel_file = format_excel(df)

    st.success("✅ Processamento concluído!")
    st.download_button(
        label="📥 Baixar Excel",
        data=excel_file,
        file_name="NotasFiscais.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
